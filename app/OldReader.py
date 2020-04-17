from openpyxl import load_workbook 
from sys import argv
from statusColor import setEmoji

def colnum(n):
   string = ""
   while n > 0:
      n, remainder = divmod(n - 1, 26)
      string = chr(65 + remainder) + string
   return string

dashboard=load_workbook(argv[1])

names=dashboard.sheetnames
output=open('index.html','w+')

### HTML
html='''
<html>
   <head>
      <meta http-equiv="content-type" content="text/html; charset=UTF-8">
      <title>CE-LAT - Project Update - Apr 3th</title>
      <link rel="stylesheet" href="style.css">
      <meta name="viewport" content="width=device-width, initial-scale=1.0"> 
   </head>
   <body>
      <div id="main">
'''
output.write(html)

### 01-NI
project="01 - NI"
ws=dashboard[project]
for i in ws.iter_rows(max_row=0):
   n_cols=len(i)
   break
n_rows=len(tuple(ws.rows))
n_cols=len(tuple(ws.columns))
rows=ws[4:n_rows-3]
project_name=project.split(' ')
output.write("\t\t<div class=\"area\">")
output.write("\t\t\t<h2>"+project_name[2]+"</h2>")
for cel in rows:
   status=cel[n_cols-1].value
   customer=cel[2].value
   name=cel[3].value
   country=cel[5].value
   if(status and status!="N/A"):
      output.write("\t\t\t<div class=\"status\">")
      output.write("\t\t\t\t<h3>"+setEmoji(cel[n_cols-1])+" "+customer+" "+country+" "+name+"</h3>")
      output.write("\t\t\t\t<ul>")
      for p in status.split('\n'):
         if (p!=' ' and p!='' and p!='\n'):
            output.write("\t\t\t\t\t<li>"+p+"</li>")
      output.write("\t\t\t\t</ul>")
      output.write("\t\t\t</div>")
output.write("\t\t</div>")

### 03-NPI
project="03 - NPI"
ws=dashboard[project]
for i in ws.iter_rows(max_row=0):
   n_cols=len(i)
   break
n_rows=len(tuple(ws.rows))
n_cols=len(tuple(ws.columns))
rows=ws[4:n_rows-3]
project_name=project.split(' ')
output.write("\t\t<div class=\"area\">")
output.write("\t\t\t<h2>"+project_name[2]+"</h2>")
for cel in rows:
   status=cel[n_cols-1].value
   customer=cel[2].value
   name=cel[3].value
   country=cel[5].value
   if(status and status!="N/A"):
      output.write("\t\t\t<div class=\"status\">")
      output.write("\t\t\t\t<h3>"+setEmoji(cel[n_cols-1])+" "+customer+" "+country+" "+name+"</h3>")
      output.write("\t\t\t\t<ul>")
      for p in status.split('\n'): 
         if (p!=' ' and p!='' and p!='\n'):
            output.write("\t\t\t\t\t<li>"+p+"</li>")
      output.write("\t\t\t\t</ul>")
      output.write("\t\t\t</div>")
output.write("\t\t</div>")

### 04-HOMOLOGATION
project="04 - HOMOLOGATION"
ws=dashboard[project]
for i in ws.iter_rows(max_row=0):
   n_cols=len(i)
   break
n_rows=len(tuple(ws.rows))
n_cols=len(tuple(ws.columns))
rows=ws[3:n_rows-2]
project_name=project.split(' ')
output.write("\t\t<div class=\"area\">")
output.write("\t\t\t<h2>"+project_name[2]+"</h2>")
for cel in rows:
   status=cel[n_cols-1].value
   customer=cel[1].value
   name=cel[5].value
   if(status and status!="N/A"):
      output.write("\t\t\t<div class=\"status\">")
      output.write("\t\t\t\t<h3>"+setEmoji(cel[n_cols-1])+" "+customer+" "+name+"</h3>")
      output.write("\t\t\t\t<ul>")
      for p in status.split('\n'):
         if (p!=' ' and p!='' and p!='\n'):
            output.write("\t\t\t\t\t<li>"+p+"</li>")
      output.write("\t\t\t\t</ul>")
      output.write("\t\t\t</div>")
output.write("\t\t</div>")

### 02-Day2
project="02 - DAY 2"
ws=dashboard[project]
for i in ws.iter_rows(max_row=0):
   n_cols=len(i)
   break
n_rows=len(tuple(ws.rows))
n_cols=len(tuple(ws.columns))
rows=ws[4:n_rows-3]
project_name=project.split(' - ')
output.write("\t\t<div class=\"area\">")
output.write("\t\t\t<h2>"+project_name[1]+"</h2>")
for cel in rows:
   status=cel[n_cols-1].value
   customer=cel[2].value
   name=cel[3].value
   country=cel[5].value
   if(status and status!="N/A"):
      output.write("\t\t\t<div class=\"status\">")
      output.write("\t\t\t\t<h3>"+setEmoji(cel[n_cols-1])+" "+customer+" "+country+" "+name+"</h3>")
      output.write("\t\t\t\t<ul>")
      for p in status.split('\n'):
         if (p!=' ' and p!='' and p!='\n'):
            output.write("\t\t\t\t\t<li>"+p+"</li>")
      output.write("\t\t\t\t</ul>")
      output.write("\t\t\t</div>")
output.write("\t\t</div>")



### /HTML
html='''
      </div>
   </body>
</html>
'''
output.write(html)

