from openpyxl import load_workbook 
from sys import argv
from statusColor import setCode

class Area:
   name=""
   projects=list()

class Project:
   name=""
   customer=""
   country=""
   code=""
   descriptions=list()

def colnum(n):
   string = ""
   while n > 0:
      n, remainder = divmod(n - 1, 26)
      string = chr(65 + remainder) + string
   return string

#Load Execel File
#excelFile=load_workbook(argv[1])
#excelFile=load_workbook('/home/andrade/2.xlsx') 
#get Sheet Names
#names=excelFile.sheetnames

def getAreas(fileName):
   excelFile=load_workbook(fileName) 
   areas=[]
   areas.append(projectNI(excelFile))
   areas.append(projectDay2(excelFile))
   areas.append(projectNPI(excelFile))
   areas.append(projectHomologations(excelFile))
   return(areas)

def projectNI(excelFile):
   tmpA=Area()
   tmpA.name="01 - NI"
   sheet=excelFile[tmpA.name]
   n_rows=len(tuple(sheet.rows))
   n_cols=len(tuple(sheet.columns))
   rows=sheet[4:n_rows]
   projects=[]
   for row in rows:
      tmpP=Project()
      tmpP.name=row[3].value
      tmpP.customer=row[2].value
      tmpP.country=row[5].value
      statusText=row[n_cols-2].value
      status=[]
      if(statusText and statusText!="N/A"):
         tmpP.code=setCode(row[n_cols-2])
         for p in statusText.split('\n'):
            if (p!=' ' and p!='' and p!='\n'):
               status.append(p)
         tmpP.descriptions=status
         projects.append(tmpP)
   tmpA.projects=projects
   return(tmpA)

def projectNPI(excelFile):
   tmpA=Area()
   tmpA.name="03 - NPI"
   sheet=excelFile[tmpA.name]
   n_rows=len(tuple(sheet.rows))
   n_cols=len(tuple(sheet.columns))
   rows=sheet[4:n_rows]
   projects=[]
   for row in rows:
      tmpP=Project()
      tmpP.name=row[3].value
      tmpP.customer=row[2].value
      tmpP.country=row[5].value
      statusText=row[n_cols-2].value
      status=[]
      if(statusText and statusText!="N/A"):
         tmpP.code=setCode(row[n_cols-2])
         for p in statusText.split('\n'):
            if (p!=' ' and p!='' and p!='\n'):
               status.append(p)
         tmpP.descriptions=status
         projects.append(tmpP)
   tmpA.projects=projects
   return(tmpA)

def projectHomologations(excelFile):
   tmpA=Area()
   tmpA.name="04 - HOMOLOGATIONS"
   sheet=excelFile[tmpA.name]
   n_rows=len(tuple(sheet.rows))
   n_cols=len(tuple(sheet.columns))
   print("rows: "+str(n_rows))
   #rows=sheet[4:n_rows-3]
   rows=sheet[4:n_rows]
   projects=[]
   for row in rows:
      tmpP=Project()
      tmpP.name=row[3].value
      tmpP.customer=row[2].value
      tmpP.country=row[5].value
      statusText=row[n_cols-2].value
      status=[]
      if(statusText and statusText!="N/A"):
         tmpP.code=setCode(row[n_cols-2])
         for p in statusText.split('\n'):
            if (p!=' ' and p!='' and p!='\n'):
               status.append(p)
         tmpP.descriptions=status
         projects.append(tmpP)
   tmpA.projects=projects
   return(tmpA)

def projectDay2(excelFile):
   tmpA=Area()
   tmpA.name="02 - DAY 2"
   sheet=excelFile[tmpA.name]
   n_rows=len(tuple(sheet.rows))
   n_cols=len(tuple(sheet.columns))
   rows=sheet[4:n_rows]
   projects=[]
   for row in rows:
      tmpP=Project()
      tmpP.name=row[3].value
      tmpP.customer=row[2].value
      tmpP.country=row[5].value
      statusText=row[n_cols-2].value
      status=[]
      if(statusText and statusText!="N/A"):
         tmpP.code=setCode(row[n_cols-2])
         for p in statusText.split('\n'):
            if (p!=' ' and p!='' and p!='\n'):
               status.append(p)
         tmpP.descriptions=status
         projects.append(tmpP)
   tmpA.projects=projects
   return(tmpA)
print("loaded")
